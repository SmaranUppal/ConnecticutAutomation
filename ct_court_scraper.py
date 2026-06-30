"""
CT Superior Court – Automated Party Name Scraper
=================================================
Searches a list of party names on the Connecticut Superior Court public
case look-up site, filters for cases filed in the last 60 days, and
exports the results to an Excel workbook.

Requirements
------------
    pip install selenium openpyxl webdriver-manager

Usage
-----
    python ct_court_scraper.py

The script opens a visible Chrome window so you can watch it work.
Set HEADLESS = True at the top if you prefer a background run.
"""

import os
import re
import time
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Configuration ─────────────────────────────────────────────────────────────

SEARCH_NAMES = [
    "J G W",
    "J.G. W",
    "J. G. W",
    "J.G.W",
    "JG W",
    "JGW",
    "J. G.W",
    "J GW",
    "Peachtree S",
    "Peach tree Settle",
    "DRB Cap",
    "Stone Street",
    "AA Ron I",
    "Abactor",
    "Abidole",
    "Adenna Med",
    "Adventura",
    "AGPI",
    "Aikman Structured Finance",
    "Annuity Transfers Ltd",
    "Apis Management",
    "Atlas Legal Funding III LP",
    "AXE Finance",
    "B.A.W.21",
    "B.R. Wright",
    "BHG Structured",
    "Bifco",
    "Blue Grape",
    "Catalina Structured Funding",
    "Concordis Group Limited",
    "Conrad Factoring",
    "Cornerstone Funding",
    "Fast Annuity S",
    "FL Assignments Corp",
    "G.D.T.R.F.B.",
    "G7 Crescenta",
    "Genex Capital Corp",
    "GJ 123",
    "Greenwood Funding",
    "Grier I",
    "Hakstol Group",
    "Hiddenview Ent, LLC",
    "JLC Capital Funding",
    "KN Direct Capital",
    "Lane Nimitz",
    "Lasko LLC",
    "Lasko, LLC",
    "Leaf 002 LLC",
    "Legere LLC",
    "Legere, LLC",
    "Lottery Funding",
    "M McDougall LLC",
    "M McDougall, LLC",
    "Majestic Funding",
    "Mic-Bry8",
    "Olive Branch Funding",
    "Palermo Group",
    "Palm Green Closing",
    "Palm Harbor",
    "Passira Mal",
    "Patriot Settlement",
    "QLS Funding",
    "Reliance Funding",
    "Rocorp Corporation",
    "RSL Funding",
    "Savannah Settlements",
    "Sempra Finance",
    "Seneca Originations",
    "SeneOne LLC",
    "Settlement Capital Corp",
    "Settlement Status",
    "Somerton LLC",
    "Somerton, LLC",
    "Stratcap Investments",
    "Stratton Asset",
    "Structured Asset",
    "TKD LLC",
    "TKD, LLC",
    "TRM V LLC",
    "TRM V, LLC",
    "Tybenz LLC",
    "Tybenz, LLC",
    "Uber Funding",
    "Vintage Equity Group",
    "Wepaymore Funding",
    "Zakho Way",
    "GREAT PLAINS MANAGEMENT CORPORATION",
    "RD FITZ LLC",
    "RD FITZ, LLC",
    "GA OFF LLC",
    "GA OFF, LLC",
    "Assured Management Corporation",
    "BENTZEN F",
]   # party last-name search strings
DAYS_BACK    = 60                              # lookback window
HEADLESS     = False                           # True = no browser window
PAGE_WAIT    = 1.5                             # seconds between page actions
DETAIL_WAIT  = 1.2                             # seconds between case-detail loads
OUTPUT_FILE  = "ct_court_results.xlsx"

BASE_URL     = "https://civilinquiry.jud.ct.gov"
SEARCH_URL   = f"{BASE_URL}/PartySearch.aspx"

# Element IDs / selectors
ID_LAST_NAME    = "ctl00_ContentPlaceHolder1_txtLastName"
ID_RADIO_PREFIX = "ctl00_ContentPlaceHolder1_rblNameSearchType"
ID_SUBMIT       = "ctl00_ContentPlaceHolder1_btnSubmit"
ID_RESULTS_TBL  = "ctl00_ContentPlaceHolder1_gvPartyResults"
ID_CASE_CAPTION = "ctl00_ContentPlaceHolder1_CaseDetailHeader1_lblCaseCaption"
RE_APPEARANCE2  = re.compile(r"lblAppearanceInfo2")

# ── Helpers ───────────────────────────────────────────────────────────────────

def cutoff_date_and_years(days: int):
    today  = datetime.today()
    cutoff = today - timedelta(days=days)
    # Collect every 2-digit year that appears in the lookback window
    years = set()
    for d in range(days + 1):
        years.add(str((cutoff + timedelta(days=d)).year)[-2:])
    return cutoff, years


def docket_year(docket_no: str) -> str | None:
    """Extract the 2-digit year from e.g. 'FBT-CV-26-1234567-S' or 'NNH-CV26-…'."""
    m = re.search(r"-CV-(\d{2})-", docket_no) or re.search(r"-CV(\d{2})-", docket_no)
    return m.group(1) if m else None


def parse_file_date(text: str) -> datetime | None:
    m = re.search(r"File Date:\s*(\d{1,2}/\d{1,2}/\d{4})", text or "")
    if m:
        try:
            return datetime.strptime(m.group(1), "%m/%d/%Y")
        except ValueError:
            pass
    return None


def parse_record_counts(driver) -> tuple[int, int, int]:
    """Return (start, end, total) from the 'Records: X-Y of Z' banner."""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        m = re.search(r"Records:\s*(\d+)-(\d+)\s+of\s+(\d+)", body)
        if m:
            return int(m.group(1)), int(m.group(2)), int(m.group(3))
    except Exception:
        pass
    return 0, 0, 0


# ── Driver setup ──────────────────────────────────────────────────────────────

def build_driver() -> webdriver.Chrome:
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--log-level=3")          # suppress console noise
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


# ── Search helpers ────────────────────────────────────────────────────────────

def wait_for_results(driver, timeout=20):
    """Wait until the results table OR any no-results indicator appears."""
    WebDriverWait(driver, timeout).until(
        lambda d: d.find_elements(By.ID, ID_RESULTS_TBL)
        or "Not Found" in d.page_source
        or "No cases" in d.page_source
        or "no records" in d.page_source.lower()
    )


def has_results(driver) -> bool:
    """Return True only if actual result rows are present (not a Not Found page)."""
    return bool(driver.find_elements(By.ID, ID_RESULTS_TBL))


def do_search(driver, name: str):
    """Navigate to search page and submit the form for `name`."""
    driver.get(SEARCH_URL)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, ID_LAST_NAME)))

    # Clear and fill last name
    field = driver.find_element(By.ID, ID_LAST_NAME)
    field.clear()
    field.send_keys(name)

    # Select "Contains" radio button
    try:
        for radio in driver.find_elements(By.NAME, "ctl00$ContentPlaceHolder1$rblNameSearchType"):
            if radio.get_attribute("value") == "Contains":
                radio.click()
                break
    except NoSuchElementException:
        pass

    # Submit
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnSubmit").click()
    wait_for_results(driver)
    time.sleep(PAGE_WAIT)


def scrape_results_page(driver) -> list[dict]:
    """Extract all data rows from the current results page."""
    rows = []
    try:
        table = driver.find_element(By.ID, ID_RESULTS_TBL)
    except NoSuchElementException:
        return rows

    for tr in table.find_elements(By.TAG_NAME, "tr"):
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) < 4:
            continue
        try:
            link_el = tds[2].find_element(By.TAG_NAME, "a")
        except NoSuchElementException:
            continue

        rows.append({
            "party_name":     tds[0].text.strip(),
            "case_name":      tds[1].text.strip(),
            "docket_no":      link_el.text.strip(),
            "docket_url":     link_el.get_attribute("href"),
            "court_location": tds[3].text.strip(),
        })
    return rows


def go_to_page(driver, page_num: int):
    """Click the GridView pager link for page_num."""
    js = (
        f"__doPostBack("
        f"'ctl00$ContentPlaceHolder1$gvPartyResults',"
        f"'Page${page_num}')"
    )
    driver.execute_script(js)
    wait_for_results(driver)
    time.sleep(PAGE_WAIT)


# ── Case detail ───────────────────────────────────────────────────────────────

def get_case_details(driver, docket_url: str) -> tuple[str, datetime | None]:
    """
    Load the case detail page and return (case_title, earliest_file_date).
    Returns ('', None) on error.
    """
    try:
        driver.get(docket_url)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, ID_CASE_CAPTION))
        )
        time.sleep(DETAIL_WAIT)

        case_title = driver.find_element(By.ID, ID_CASE_CAPTION).text.strip()

        # Collect all appearance-info-2 labels (file dates)
        dates = []
        for el in driver.find_elements(By.XPATH, "//*[contains(@id,'lblAppearanceInfo2')]"):
            d = parse_file_date(el.text)
            if d:
                dates.append(d)

        return case_title, min(dates) if dates else None

    except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
        print(f"      ⚠  Could not load {docket_url}: {e}")
        return "", None


# ── Main scraper ──────────────────────────────────────────────────────────────

def scrape() -> list[dict]:
    cutoff, valid_years = cutoff_date_and_years(DAYS_BACK)
    today = datetime.today()

    print(f"\n{'='*60}")
    print(f"  CT Superior Court Scraper")
    print(f"  Today  : {today.strftime('%m/%d/%Y')}")
    print(f"  Cutoff : {cutoff.strftime('%m/%d/%Y')}  (last {DAYS_BACK} days)")
    print(f"  Years  : {sorted(valid_years)}")
    print(f"  Names  : {SEARCH_NAMES}")
    print(f"{'='*60}\n")

    driver = build_driver()
    all_results = []

    try:
        for search_name in SEARCH_NAMES:
            print(f"🔍  Searching: '{search_name}'")
            do_search(driver, search_name)

            if not has_results(driver):
                print(f"    No records found.\n")
                continue

            start, end, total = parse_record_counts(driver)
            print(f"    {total} total record(s) found.")

            # ── Collect rows across all pages ──────────────────────────────
            all_rows: list[dict] = []
            page = 1

            while True:
                rows = scrape_results_page(driver)
                all_rows.extend(rows)
                _, end, _ = parse_record_counts(driver)
                print(f"    Page {page}: {len(rows)} rows  (cumulative: {len(all_rows)} / {total})")

                if end >= total:
                    break                       # no more pages
                page += 1
                go_to_page(driver, page)

            # ── Filter by docket year ──────────────────────────────────────
            year_candidates = [
                r for r in all_rows
                if docket_year(r["docket_no"]) in valid_years
            ]
            print(f"    After year filter {sorted(valid_years)}: "
                  f"{len(year_candidates)} candidate(s)")

            # ── Load each candidate and check the file date ────────────────
            for i, row in enumerate(year_candidates, 1):
                print(f"    [{i}/{len(year_candidates)}] {row['docket_no']} ...", end=" ")
                case_title, file_date = get_case_details(driver, row["docket_url"])

                if file_date is None:
                    print("no file date — skip")
                    continue
                if file_date < cutoff:
                    print(f"filed {file_date.strftime('%m/%d/%Y')} — before cutoff, skip")
                    continue

                print(f"✅  filed {file_date.strftime('%m/%d/%Y')} — IN RANGE")
                all_results.append({
                    "search_name":    search_name,
                    "case_title":     case_title or row["case_name"],
                    "docket_no":      row["docket_no"],
                    "court_location": row["court_location"],
                    "file_date":      file_date.strftime("%m/%d/%Y"),
                    "docket_url":     row["docket_url"],
                })

            print()

    finally:
        driver.quit()

    return all_results


# ── Excel export ──────────────────────────────────────────────────────────────

HEADERS    = ["Search Term", "Case Title", "Docket No.", "Court Location",
              "File Date", "Case URL"]
COL_WIDTHS = [14, 52, 24, 20, 14, 60]
# Column index (1-based) of the docket number — used for duplicate detection
DOCKET_COL = HEADERS.index("Docket No.") + 1   # → 3


def _styles():
    thin = Side(style="thin", color="B8B8B8")
    return {
        "hdr_font":  Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "hdr_fill":  PatternFill("solid", start_color="1F3864"),
        "data_font": Font(name="Arial", size=10),
        "alt_fill":  PatternFill("solid", start_color="DCE6F1"),
        "center":    Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left":      Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "border":    Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _write_header(ws, st):
    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = st["hdr_font"]; c.fill = st["hdr_fill"]
        c.alignment = st["center"]; c.border = st["border"]
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "A2"


def _append_rows(ws, rows: list[dict], st, start_row: int):
    """Write `rows` to `ws` beginning at `start_row`, alternating fill by absolute row."""
    for offset, r in enumerate(rows):
        row_idx = start_row + offset
        fill = st["alt_fill"] if row_idx % 2 == 0 else None
        values = [r["search_name"], r["case_title"], r["docket_no"],
                  r["court_location"], r["file_date"], r["docket_url"]]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = st["data_font"]; c.border = st["border"]
            c.alignment = st["left"] if col == 2 else st["center"]
            if fill:
                c.fill = fill
        ws.row_dimensions[row_idx].height = 18


def _existing_dockets(ws) -> set[str]:
    """Return the set of docket numbers already in the sheet (skips header row)."""
    dockets = set()
    for row in ws.iter_rows(min_row=2, min_col=DOCKET_COL, max_col=DOCKET_COL, values_only=True):
        val = row[0]
        if val:
            dockets.add(str(val).strip())
    return dockets


def _update_run_info(wb, new_count: int, skipped: int):
    """Overwrite (or create) the Run Info sheet with stats from this run."""
    if "Run Info" in wb.sheetnames:
        del wb["Run Info"]
    mi = wb.create_sheet("Run Info")
    mi.column_dimensions["A"].width = 24
    mi.column_dimensions["B"].width = 42
    bold = Font(name="Arial", size=10, bold=True)
    norm = Font(name="Arial", size=10)
    cutoff, _ = cutoff_date_and_years(DAYS_BACK)
    meta = [
        ("Last Run Date",      datetime.today().strftime("%m/%d/%Y %H:%M")),
        ("Lookback (days)",    str(DAYS_BACK)),
        ("Cutoff Date",        cutoff.strftime("%m/%d/%Y")),
        ("Names Searched",     ", ".join(SEARCH_NAMES)),
        ("New Cases Added",    str(new_count)),
        ("Duplicates Skipped", str(skipped)),
    ]
    for r_idx, (k, v) in enumerate(meta, 1):
        mi.cell(r_idx, 1, k).font = bold
        mi.cell(r_idx, 2, v).font = norm


def export_excel(results: list[dict], output_path: str):
    st = _styles()

    # ── Deduplicate within this run first ──────────────────────────────────
    seen_this_run: set[str] = set()
    unique_results: list[dict] = []
    for r in results:
        key = r["docket_no"].strip()
        if key not in seen_this_run:
            seen_this_run.add(key)
            unique_results.append(r)
        else:
            print(f"  ⚠  Duplicate within run, skipping: {key}")

    intra_dupes = len(results) - len(unique_results)

    # ── Load or create workbook ────────────────────────────────────────────
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        if "CT Court Cases" in wb.sheetnames:
            ws = wb["CT Court Cases"]
        else:
            ws = wb.create_sheet("CT Court Cases", 0)
            _write_header(ws, st)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "CT Court Cases"
        _write_header(ws, st)

    # ── Filter out dockets already in the sheet ────────────────────────────
    existing = _existing_dockets(ws)
    net_new = [r for r in unique_results if r["docket_no"].strip() not in existing]
    inter_dupes = len(unique_results) - len(net_new)

    for r in unique_results:
        if r["docket_no"].strip() in existing:
            print(f"  ⚠  Already in sheet, skipping: {r['docket_no']}")

    # ── Append net-new rows ────────────────────────────────────────────────
    next_row = ws.max_row + 1 if ws.max_row and ws.max_row > 1 else 2
    _append_rows(ws, net_new, st, next_row)

    total_skipped = intra_dupes + inter_dupes
    _update_run_info(wb, len(net_new), total_skipped)

    wb.save(output_path)
    print(f"\n✅  {len(net_new)} new row(s) added  |  "
          f"{intra_dupes} intra-run duplicate(s)  |  "
          f"{inter_dupes} already-in-sheet duplicate(s)  →  {output_path}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    results = scrape()
    export_excel(results, OUTPUT_FILE)